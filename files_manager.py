import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill

green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

# Исходный список словарей
def into_csv_data(data, csv_file):
    with open(csv_file, mode='w', newline='', encoding='utf-8') as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=["name",
                                                      'last_match',
                                                      'matches_played',
                                                      'goals',
                                                      'assists',
                                                      'points',
                                                      'season_name'])
        writer.writeheader()
        writer.writerows(data)


def into_excel_data(data, excel_file):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Player Stats"

    # Заголовки
    headers = ["name", "last_match", "matches_played", "goals", "assists", "points", "season_name"]
    sheet.append(headers)

    # Добавляем данные и окрашиваем строки
    for row in data:
        try:
            res = int(row.get('points')) / row.get('matches_played')

        except:
            res = 0

        row_data = [row.get(key, "") for key in headers]
        if row.get('last_match') == 'не заявлен' and res > 0.5:
            fill = orange_fill

        elif row.get('last_match') == 'не заявлен':
            fill = red_fill

        elif res > 0.5:
            fill = green_fill
        else:
            fill = None

        # Добавляем данные в строку и применяем окраску ко всей строке
        row_idx = sheet.max_row + 1
        for idx, cell in enumerate(row_data, start=1):
            cell_obj = sheet.cell(row=row_idx, column=idx, value=cell)
            if fill:
                cell_obj.fill = fill  # Применяем заливку

    # Сохраняем файл
    workbook.save(excel_file)
